"""Write outputs: JSON, history CSV, a formula-driven Excel model, dashboard data."""
import csv, json, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

BLUE = "1F4E79"
HEADER = PatternFill("solid", fgColor=BLUE)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ARIAL = "Arial"


def _ts():
    return datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")


def write_json(result, narrative, date, path):
    payload = {
        "date": date, "generated_at": _ts(),
        "aggregates": result["aggregates"], "narrative": narrative,
        "stories": result["stories"],
    }
    if result.get("agent_review"):
        payload["agent_review"] = result["agent_review"]
    if result.get("data_health"):
        payload["data_health"] = result["data_health"]
    if result.get("aggregates_trend"):
        payload["aggregates_trend"] = result["aggregates_trend"]
    if result.get("sets"):
        payload["sets"] = result["sets"]
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(payload, f, indent=2)
    return payload


def append_history(result, date, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    new = not os.path.exists(path)
    with open(path, "a", newline="") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["date", "story_id", "name", "set", "level", "band", "raw_level"])
        for s in result["stories"]:
            w.writerow([date, s["id"], s["name"], s["set"], s["level"], s["band"], s["raw_level"]])


def _hdr(ws, row, headers, widths):
    for c, (h, wd) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=ARIAL, bold=True, color="FFFFFF")
        cell.fill = HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = wd


def build_workbook(config, result, narrative, date, history_path, path):
    wb = Workbook()

    # ---- Indicators sheet (inputs + band/delta formulas) ----
    ind = wb.active
    ind.title = "Indicators"
    _hdr(ind, 1, ["Story", "Story ID", "Indicator", "Value", "Direction", "Good",
                  "Warn", "Baseline band", "Current band", "Weight", "Delta",
                  "Source", "Cadence", "Last updated", "Note"],
         [22, 16, 34, 11, 13, 8, 8, 12, 12, 8, 8, 22, 11, 14, 30])
    r = 2
    for s in config["stories"]:
        sc = next(x for x in result["stories"] if x["id"] == s["id"])
        det = {d["id"]: d for d in sc["indicators"]}
        for indic in s["indicators"]:
            d = det[indic["id"]]
            ind.cell(r, 1, s["name"]).font = Font(name=ARIAL)
            ind.cell(r, 2, s["id"]).font = Font(name=ARIAL)
            ind.cell(r, 3, indic["label"]).font = Font(name=ARIAL)
            vcell = ind.cell(r, 4, d["value"])
            vcell.font = Font(name=ARIAL, color="0000FF")  # blue = editable input
            ind.cell(r, 5, indic["direction"]).font = Font(name=ARIAL)
            ind.cell(r, 6, indic["good"]).font = Font(name=ARIAL)
            ind.cell(r, 7, indic["warn"]).font = Font(name=ARIAL)
            ind.cell(r, 8, indic.get("baseline_band", 0)).font = Font(name=ARIAL)
            if indic["direction"] == "higher_worse":
                band_f = f"=IF(D{r}<=F{r},0,IF(D{r}<=G{r},1,2))"
            else:
                band_f = f"=IF(D{r}>=F{r},0,IF(D{r}>=G{r},1,2))"
            ind.cell(r, 9, band_f).font = Font(name=ARIAL)
            ind.cell(r, 10, indic["weight"]).font = Font(name=ARIAL)
            ind.cell(r, 11, f"=J{r}*(I{r}-H{r})/2").font = Font(name=ARIAL)
            ind.cell(r, 12, indic["source"]).font = Font(name=ARIAL, size=9)
            ind.cell(r, 13, indic["cadence"]).font = Font(name=ARIAL)
            ind.cell(r, 14, "live" if not d["stale"] else "baseline").font = Font(name=ARIAL)
            ind.cell(r, 15, indic.get("note", "")).font = Font(name=ARIAL, size=9)
            for c in range(1, 16):
                ind.cell(r, c).border = BORDER
            r += 1
    ind.freeze_panes = "A2"

    # ---- Dashboard sheet (pulls from Indicators) ----
    dash = wb.create_sheet("Dashboard", 0)
    dash["A1"] = "GLOBAL MACRO RISK TRACKER"
    dash["A1"].font = Font(name=ARIAL, bold=True, size=16, color=BLUE)
    dash["A2"] = f"As of {date}  -  generated {_ts()}  -  levels recalculated from the Indicators tab"
    dash["A2"].font = Font(name=ARIAL, italic=True, color="595959")
    _hdr(dash, 4, ["Story", "Story ID", "Set", "Base", "Indicator adj.", "Level (1-10)", "Band"],
         [34, 16, 6, 7, 13, 12, 12])
    start = 5
    for i, s in enumerate(config["stories"]):
        rr = start + i
        dash.cell(rr, 1, s["name"]).font = Font(name=ARIAL)
        dash.cell(rr, 2, s["id"]).font = Font(name=ARIAL)
        dash.cell(rr, 3, s["set"]).font = Font(name=ARIAL)
        dash.cell(rr, 4, s["base_level"]).font = Font(name=ARIAL)
        dash.cell(rr, 5, f"=ROUND(SUMIF(Indicators!$B:$B,B{rr},Indicators!$K:$K),2)").font = Font(name=ARIAL)
        dash.cell(rr, 6, f"=MEDIAN(1,10,ROUND(D{rr}+E{rr},0))").font = Font(name=ARIAL, bold=True)
        dash.cell(rr, 7, f'=IF(F{rr}<=2,"Low",IF(F{rr}<=4,"Moderate",IF(F{rr}<=6,"Elevated",IF(F{rr}<=8,"High","Severe"))))').font = Font(name=ARIAL)
        for c in range(1, 8):
            dash.cell(rr, c).border = BORDER
    last = start + len(config["stories"]) - 1
    # aggregates (one row per set, dynamic)
    agg_row = last + 2
    set_nums = sorted({s["set"] for s in config["stories"]})
    set_names = (config.get("sets") or {})
    for k, sn in enumerate(set_nums):
        nm = (set_names.get(sn) or {}).get("name", f"Set {sn}")
        dash.cell(agg_row + k, 1, f"Aggregate - {nm}").font = Font(name=ARIAL, bold=True)
        dash.cell(agg_row + k, 6,
                  f"=ROUND(AVERAGEIF($C${start}:$C${last},{sn},$F${start}:$F${last}),1)"
                  ).font = Font(name=ARIAL, bold=True)
    orow = agg_row + len(set_nums)
    dash.cell(orow, 1, "Aggregate - Overall").font = Font(name=ARIAL, bold=True)
    dash.cell(orow, 6, f"=ROUND(AVERAGE($F${start}:$F${last}),1)").font = Font(name=ARIAL, bold=True)
    dash.conditional_formatting.add(
        f"F{start}:F{last}",
        ColorScaleRule(start_type="num", start_value=1, start_color="2E7D32",
                       mid_type="num", mid_value=5, mid_color="F5A623",
                       end_type="num", end_value=10, end_color="C0392B"))
    dash.freeze_panes = "A5"

    # ---- Daily Log (rebuilt from history.csv) ----
    log = wb.create_sheet("Daily Log")
    _hdr(log, 1, ["Date", "Story ID", "Name", "Set", "Level", "Band", "Raw"],
         [12, 16, 30, 6, 8, 11, 8])
    if os.path.exists(history_path):
        with open(history_path) as f:
            rows = list(csv.reader(f))[1:]
        for i, row in enumerate(rows[-1000:], start=2):
            for c, val in enumerate(row, 1):
                log.cell(i, c, val).font = Font(name=ARIAL)
    log.freeze_panes = "A2"

    # ---- Methodology + narrative ----
    meth = wb.create_sheet("Methodology")
    meth.column_dimensions["A"].width = 110
    lines = [
        ("GLOBAL MACRO RISK TRACKER - methodology", True),
        ("", False),
        ("Each story has a human-set base_level (1-10). Each indicator is classified into a band:", False),
        ("0 = benign, 1 = watch, 2 = stress, using its direction and the Good/Warn thresholds.", False),
        ("delta = weight * (current_band - baseline_band) / 2.  Level = clamp(base_level + sum(deltas), 1, 10).", False),
        ("Live data nudges the analyst anchor; every move is visible on the Indicators tab.", False),
        ("Blue cells on the Indicators tab are editable inputs (update them, or let the script do it).", False),
        ("", False),
        ("Sources: fred:CODE / fred_spread:A:B (keyless FRED CSV), market:SYMBOL (Yahoo), manual (judgement / slow data).", False),
        ("Cadence varies: market indicators move daily; fundamentals are monthly/quarterly; policy items are event-driven.", False),
        ("This is informational analysis, not investment, legal, or financial advice.", False),
        ("", False),
        (f"--- Daily narrative ({date}) ---", True),
    ]
    rr = 1
    for txt, bold in lines:
        c = meth.cell(rr, 1, txt)
        c.font = Font(name=ARIAL, bold=bold, color=BLUE if bold else "000000")
        rr += 1
    for line in narrative.split("\n"):
        meth.cell(rr, 1, line).font = Font(name=ARIAL)
        meth.cell(rr, 1).alignment = Alignment(wrap_text=True)
        rr += 1

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path
